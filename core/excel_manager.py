# excel_manager.py
import os
import json
from pathlib import Path
import openpyxl  # Para validar y abrir Excel

class ExcelManager:
    def __init__(self, config_path: Path):
        self.config_path = config_path

    def load_excels(self):
        if not self.config_path.exists():
            return []
        with open(self.config_path, "r", encoding="utf-8") as f:
            return json.load(f).get("excels", [])

    def validate_excel(self, path):
        """
        Validar que el archivo exista, se pueda abrir y tenga formato correcto
        """
        if not os.path.exists(path):
            return False
        try:
            wb = openpyxl.load_workbook(path)
            # Aquí puedes validar hojas, columnas, etc.
            wb.close()
            return True
        except:
            return False

    def update_excel(self, path, backup_path=None):
        """
        Placeholder: implementar la actualización real
        """
        # Ejemplo: crear backup
        if backup_path:
            import shutil
            shutil.copy(path, backup_path)

    def validate_rows(self, path, min_rows=1):
        """
        Validar que el Excel tenga al menos 'min_rows'
        """
        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        rows = ws.max_row
        wb.close()
        return rows >= min_rows
